# utils.py
"""
Fonctions utilitaires pour l'application Finance Viewer.
"""

import re
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def clean_text(text):
    """
    Nettoie un texte pour le rendre utilisable comme nom de fichier ou de feuille Excel.
    Remplace les caractères spéciaux problématiques.

    Args:
        text (str): Texte à nettoyer

    Returns:
        str: Texte nettoyé
    """
    clean_name = re.sub(r'[\\/*?:"<>|=]', '_', str(text))
    return clean_name


def create_excel(data, sheet_name="Data"):
    """
    Crée un fichier Excel à partir des données financières.

    Args:
        data (DataFrame): Données à exporter
        sheet_name (str): Nom de la feuille Excel

    Returns:
        bytes: Contenu du fichier Excel
    """
    output = io.BytesIO()

    # Formater les données pour l'export
    export_data = data.copy()

    # Calculer la variation quotidienne
    export_data['Variation (%)'] = export_data['Close'].pct_change() * 100

    # Convertir les indices en dates au format YYYY-MM-DD
    export_data.index = [d.strftime('%Y-%m-%d') for d in export_data.index]
    export_data.index.name = 'Date'

    # Réorganiser les colonnes
    export_data = export_data[['Close', 'High', 'Low', 'Open', 'Variation (%)', 'Volume']]

    # Renommer les colonnes
    export_data.columns = ['Price', 'High', 'Low', 'Open', 'Variation (%)', 'Volume']

    # Nettoyer le nom de la feuille
    clean_sheet_name = clean_text(sheet_name)
    # Limiter la longueur à 31 caractères (limite Excel)
    if len(clean_sheet_name) > 31:
        clean_sheet_name = clean_sheet_name[:31]

    # Créer et configurer manuellement le fichier Excel
    workbook = Workbook()
    ws = workbook.active
    ws.title = clean_sheet_name

    # Ajouter la ligne d'en-tête
    headers = ['Date'] + list(export_data.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Ajouter les données
    for row_idx, (date_idx, row_data) in enumerate(export_data.iterrows(), start=2):
        # Ajouter la date
        ws.cell(row=row_idx, column=1, value=date_idx)

        # Ajouter les autres colonnes
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Pour la colonne Variation (%), appliquer le format pourcentage
            if col_idx == 6:  # 6 est l'index pour Variation (%)
                if pd.notna(value):
                    cell.value = value / 100  # Convertir en décimal pour Excel
                    cell.number_format = '0.00%'  # Format pourcentage avec 2 décimales

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    workbook.save(output)
    processed_data = output.getvalue()
    return processed_data


def format_volume(volume):
    """
    Formate un volume pour l'affichage (k, M, G).

    Args:
        volume (float): Volume à formater

    Returns:
        str: Volume formaté
    """
    vol = float(volume)
    if vol >= 1e9:
        return f"{vol / 1e9:.2f} G"
    elif vol >= 1e6:
        return f"{vol / 1e6:.2f} M"
    elif vol >= 1e3:
        return f"{vol / 1e3:.2f} k"
    else:
        return f"{vol:.2f}"
